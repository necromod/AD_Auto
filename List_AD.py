from datetime import datetime, timedelta
from ldap3 import Server, Connection, NTLM, ALL, MODIFY_REPLACE, Tls
import os
import getpass
import socket
import ssl
import sys
import subprocess

# Tenta importar openpyxl e instala se necessário
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_DISPONIVEL = True
except ImportError:
    OPENPYXL_DISPONIVEL = False
    print("⚠ Biblioteca openpyxl não encontrada. Instalando...")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        OPENPYXL_DISPONIVEL = True
        print("✓ Biblioteca openpyxl instalada com sucesso!")
    except Exception as e:
        print(f"❌ Erro ao instalar openpyxl: {e}")
        print("A função de relatório em Excel não estará disponível.")
        OPENPYXL_DISPONIVEL = False

# Obtém o nome do usuário logado no sistema
def get_usuario_logado():
    return os.getlogin()

def _testar_conectividade_rede(dominio_dns):
    """Testa conectividade de rede básica com o servidor AD"""
    
    print(f"  🌐 Testando resolução DNS de {dominio_dns}...")
    try:
        endereco_ip = socket.gethostbyname(dominio_dns)
        print(f"    ✓ DNS resolvido: {dominio_dns} → {endereco_ip}")
    except socket.gaierror as e:
        print(f"    ✗ Falha na resolução DNS: {e}")
        return False
    
    print(f"  🔌 Testando conectividade TCP...")
    portas_teste = [389, 636, 3268, 3269]
    for porta in portas_teste:
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(3)
            resultado = sock.connect_ex((endereco_ip, porta))
            sock.close()
            
            if resultado == 0:
                print(f"    ✓ Porta {porta} acessível")
            else:
                print(f"    ✗ Porta {porta} inacessível")
                
        except Exception as e:
            print(f"    ✗ Erro testando porta {porta}: {e}")
    
    return True

# Obtém a conexão com o Active Directory baseado no usuário logado
def get_conexao():
    usuario = get_usuario_logado()
    dominio_netbios = os.environ.get('USERDOMAIN')
    dominio_dns = os.environ.get('USERDNSDOMAIN')
    usuario_completo = f'{dominio_netbios}\\{usuario}'
    
    # Verifica se o domínio DNS foi obtido corretamente
    if not dominio_dns:
        raise Exception("Não foi possível obter o domínio DNS.")

    print("🔍 CONECTANDO AO ACTIVE DIRECTORY")
    print(f"  Usuário: {usuario_completo}")
    print(f"  Domínio: {dominio_dns}")
    
    # Testa conectividade básica primeiro
    print("\n📡 Testando conectividade...")
    _testar_conectividade_rede(dominio_dns)
    
    print("\n🔐 Estabelecendo conexão...")
    
    # Lista de configurações para tentar (sem validação TLS)
    configuracoes = [
        # Configuração 1: LDAP simples na porta 389
        {
            'tls': None,
            'porta': 389,
            'use_ssl': False,
            'descricao': "LDAP porta 389 (simples)"
        },
        # Configuração 2: LDAPS sem validação
        {
            'tls': Tls(validate=ssl.CERT_NONE),
            'porta': 636,
            'use_ssl': True,
            'descricao': "LDAPS porta 636 (sem validação)"
        }
    ]
    
    # Tenta cada configuração
    for i, config in enumerate(configuracoes, 1):
        try:
            print(f"  Tentativa {i}/{len(configuracoes)}: {config['descricao']}")
            
            servidor = Server(dominio_dns, port=config['porta'], use_ssl=config['use_ssl'], tls=config['tls'], get_info=ALL)
            conexao = _tentar_conexao(servidor, usuario_completo, config)
            
            if conexao:
                return conexao
                
        except Exception as e:
            print(f"  ✗ Falhou: {e}")
            continue
    
    print("\n❌ ERRO: Não foi possível conectar ao Active Directory")
    print("Verifique as configurações de rede e SSL do servidor AD.")
    sys.exit(1)

def _tentar_conexao(servidor, usuario_completo, config):
    tentativas = 3
    for tentativa in range(tentativas):
        try:
            # Solicita a senha do usuário logado
            if tentativa == 0:
                senha = getpass.getpass("Digite sua senha do AD: ")
            else:
                print(f"Tentativa {tentativa + 1} de {tentativas}")
                senha = getpass.getpass("Digite sua senha do AD: ")
                
            # Conecta ao AD
            conexao = Connection(
                servidor, 
                user=usuario_completo, 
                password=senha, 
                authentication=NTLM, 
                auto_bind=True,
                receive_timeout=10
            )
            
            # Verifica se precisa fazer StartTLS
            if config.get('start_tls', False):
                if not conexao.start_tls():
                    raise Exception("Falha ao iniciar StartTLS")
                print(f"✓ Conexão StartTLS estabelecida - {config['descricao']}")
            else:
                print(f"✓ Conexão estabelecida - {config['descricao']}")
            
            # Teste básico de conectividade
            try:
                base_dn = conexao.server.info.other['defaultNamingContext'][0]
                print(f"✓ Conectado ao domínio: {base_dn}")
            except Exception as teste_erro:
                print(f"⚠ Aviso: Conexão estabelecida mas teste falhou: {teste_erro}")
            
            return conexao
            
        except Exception as e:
            erro_str = str(e)
            if "invalidCredentials" in erro_str or "49" in erro_str:
                print("✗ Credenciais incorretas.")
                if tentativa < tentativas - 1:
                    continue
                else:
                    print("Número máximo de tentativas excedido.")
                    return None
            else:
                raise e
    
    return None

# Obtém o DN do server
def get_base_dn(conexao):
    return conexao.server.info.other['defaultNamingContext'][0]

def gerar_contas_ativas(conexao):
    """Gera relatório com todas as contas ativas"""
    if not OPENPYXL_DISPONIVEL:
        print("❌ ERRO: Biblioteca openpyxl não está disponível.")
        return
    
    base_dn = get_base_dn(conexao)
    print("\n📊 GERANDO RELATÓRIO - CONTAS ATIVAS")
    
    # Busca todos os usuários ativos
    filtro = '(&(objectClass=user)(!(sAMAccountName=*$))(!(userAccountControl:1.2.840.113556.1.4.803:=2)))'
    atributos = ['sAMAccountName', 'displayName', 'title', 'mail', 'whenCreated', 'lastLogon', 'lastLogonTimestamp']
    
    dados_usuarios = buscar_usuarios_com_paginacao(conexao, base_dn, filtro, atributos)
    
    if not dados_usuarios:
        print("❌ Nenhuma conta ativa encontrada.")
        return
    
    # Processa os dados
    usuarios_processados = []
    for entry in dados_usuarios:
        login = entry.sAMAccountName.value if hasattr(entry, 'sAMAccountName') and entry.sAMAccountName.value else 'N/A'
        nome = entry.displayName.value if hasattr(entry, 'displayName') and entry.displayName.value else login
        cargo = entry.title.value if hasattr(entry, 'title') and entry.title.value else 'Não informado'
        email = entry.mail.value if hasattr(entry, 'mail') and entry.mail.value else 'N/A'
        
        # Data de criação
        data_criacao = 'N/A'
        if hasattr(entry, 'whenCreated') and entry.whenCreated.value:
            try:
                created_dt = entry.whenCreated.value.replace(tzinfo=None)
                data_criacao = created_dt.strftime('%d/%m/%Y')
            except:
                data_criacao = 'Erro ao converter'
        
        # Data de último logon
        data_ultimo_logon = 'Nunca'
        if hasattr(entry, 'lastLogon') and entry.lastLogon.value:
            try:
                logon_dt = entry.lastLogon.value.replace(tzinfo=None)
                data_ultimo_logon = logon_dt.strftime('%d/%m/%Y')
            except:
                pass
        
        usuarios_processados.append({
            'Login': login,
            'Nome': nome,
            'E-mail': email,
            'Cargo': cargo,
            'Status': 'Ativo',
            'Data de Criação': data_criacao,
            'Último Logon': data_ultimo_logon
        })
    
    # Ordena por nome
    usuarios_processados.sort(key=lambda x: x['Nome'] or 'ZZZ')
    
    # Gera a planilha
    nome_arquivo = f"Contas_Ativas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    colunas = ['Login', 'Nome', 'E-mail', 'Cargo', 'Status', 'Data de Criação', 'Último Logon']
    gerar_planilha(usuarios_processados, nome_arquivo, "CONTAS ATIVAS", colunas)

def gerar_contas_desabilitadas_desde_abril(conexao):
    """Gera relatório com contas desabilitadas a partir de 01/04/2024"""
    if not OPENPYXL_DISPONIVEL:
        print("❌ ERRO: Biblioteca openpyxl não está disponível.")
        return
    
    base_dn = get_base_dn(conexao)
    print("\n📊 GERANDO RELATÓRIO - CONTAS DESABILITADAS A PARTIR DE 01/04/2024")
    
    data_corte = datetime(2024, 4, 1)
    
    # Busca todos os usuários desabilitados
    filtro = '(&(objectClass=user)(!(sAMAccountName=*$))(userAccountControl:1.2.840.113556.1.4.803:=2))'
    atributos = ['sAMAccountName', 'displayName', 'title', 'mail', 'whenCreated', 'lastLogon', 'lastLogonTimestamp', 'accountExpires']
    
    dados_usuarios = buscar_usuarios_com_paginacao(conexao, base_dn, filtro, atributos)
    
    if not dados_usuarios:
        print("❌ Nenhuma conta desabilitada encontrada.")
        return
    
    # Processa os dados e aplica filtro de data
    usuarios_processados = []
    for entry in dados_usuarios:
        # Verifica se foi desabilitado a partir de abril/2024
        # (Para isso, vamos usar a lógica de que deve ter tido logon em 2024)
        data_ultimo_logon = None
        
        if hasattr(entry, 'lastLogon') and entry.lastLogon.value:
            try:
                data_ultimo_logon = entry.lastLogon.value.replace(tzinfo=None)
            except:
                pass
        
        if hasattr(entry, 'lastLogonTimestamp') and entry.lastLogonTimestamp.value:
            try:
                logon_timestamp = entry.lastLogonTimestamp.value.replace(tzinfo=None)
                if data_ultimo_logon:
                    data_ultimo_logon = max(data_ultimo_logon, logon_timestamp)
                else:
                    data_ultimo_logon = logon_timestamp
            except:
                pass
        
        # Só inclui se teve logon a partir de abril/2024
        if data_ultimo_logon and data_ultimo_logon >= data_corte:
            login = entry.sAMAccountName.value if hasattr(entry, 'sAMAccountName') and entry.sAMAccountName.value else 'N/A'
            nome = entry.displayName.value if hasattr(entry, 'displayName') and entry.displayName.value else login
            cargo = entry.title.value if hasattr(entry, 'title') and entry.title.value else 'Não informado'
            email = entry.mail.value if hasattr(entry, 'mail') and entry.mail.value else 'N/A'
            
            # Data de criação
            data_criacao = 'N/A'
            if hasattr(entry, 'whenCreated') and entry.whenCreated.value:
                try:
                    created_dt = entry.whenCreated.value.replace(tzinfo=None)
                    data_criacao = created_dt.strftime('%d/%m/%Y')
                except:
                    data_criacao = 'Erro ao converter'
            
            usuarios_processados.append({
                'Login': login,
                'Nome': nome,
                'E-mail': email,
                'Cargo': cargo,
                'Status': 'Inativo',
                'Data de Criação': data_criacao,
                'Último Logon': data_ultimo_logon.strftime('%d/%m/%Y')
            })
    
    # Ordena por nome
    usuarios_processados.sort(key=lambda x: x['Nome'] or 'ZZZ')
    
    # Gera a planilha
    nome_arquivo = f"Contas_Desabilitadas_Desde_Abril_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    colunas = ['Login', 'Nome', 'E-mail', 'Cargo', 'Status', 'Data de Criação', 'Último Logon']
    gerar_planilha(usuarios_processados, nome_arquivo, "CONTAS DESABILITADAS A PARTIR DE 01/04/2024", colunas)

def gerar_contas_criadas_em_2024(conexao):
    """Gera relatório com contas criadas somente em 2024"""
    if not OPENPYXL_DISPONIVEL:
        print("❌ ERRO: Biblioteca openpyxl não está disponível.")
        return
    
    base_dn = get_base_dn(conexao)
    print("\n📊 GERANDO RELATÓRIO - CONTAS CRIADAS EM 2024")
    
    data_inicio_2024 = datetime(2024, 1, 1)
    data_fim_2024 = datetime(2025, 1, 1)
    
    # Busca todos os usuários
    filtro = '(&(objectClass=user)(!(sAMAccountName=*$)))'
    atributos = ['sAMAccountName', 'displayName', 'title', 'mail', 'whenCreated', 'userAccountControl', 'lastLogon', 'lastLogonTimestamp']
    
    dados_usuarios = buscar_usuarios_com_paginacao(conexao, base_dn, filtro, atributos)
    
    if not dados_usuarios:
        print("❌ Nenhum usuário encontrado.")
        return
    
    # Processa os dados e aplica filtro de data de criação
    usuarios_processados = []
    for entry in dados_usuarios:
        # Verifica data de criação
        created_dt = None
        if hasattr(entry, 'whenCreated') and entry.whenCreated.value:
            try:
                created_dt = entry.whenCreated.value.replace(tzinfo=None)
            except:
                continue
        
        # Só inclui se foi criado em 2024
        if created_dt and data_inicio_2024 <= created_dt < data_fim_2024:
            login = entry.sAMAccountName.value if hasattr(entry, 'sAMAccountName') and entry.sAMAccountName.value else 'N/A'
            nome = entry.displayName.value if hasattr(entry, 'displayName') and entry.displayName.value else login
            cargo = entry.title.value if hasattr(entry, 'title') and entry.title.value else 'Não informado'
            email = entry.mail.value if hasattr(entry, 'mail') and entry.mail.value else 'N/A'
            
            # Status da conta
            user_account_control = int(entry.userAccountControl.value) if hasattr(entry, 'userAccountControl') and entry.userAccountControl.value else 0
            is_disabled = bool(user_account_control & 0x0002)
            status = 'Inativo' if is_disabled else 'Ativo'
            
            # Data de último logon
            data_ultimo_logon = 'Nunca'
            if hasattr(entry, 'lastLogon') and entry.lastLogon.value:
                try:
                    logon_dt = entry.lastLogon.value.replace(tzinfo=None)
                    data_ultimo_logon = logon_dt.strftime('%d/%m/%Y')
                except:
                    pass
            
            usuarios_processados.append({
                'Login': login,
                'Nome': nome,
                'E-mail': email,
                'Cargo': cargo,
                'Status': status,
                'Data de Criação': created_dt.strftime('%d/%m/%Y'),
                'Último Logon': data_ultimo_logon
            })
    
    # Ordena por nome
    usuarios_processados.sort(key=lambda x: x['Nome'] or 'ZZZ')
    
    # Gera a planilha
    nome_arquivo = f"Contas_Criadas_2024_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    colunas = ['Login', 'Nome', 'E-mail', 'Cargo', 'Status', 'Data de Criação', 'Último Logon']
    gerar_planilha(usuarios_processados, nome_arquivo, "CONTAS CRIADAS EM 2024", colunas)

def gerar_contas_desabilitadas_em_2024(conexao):
    """Gera relatório com contas desabilitadas somente em 2024"""
    if not OPENPYXL_DISPONIVEL:
        print("❌ ERRO: Biblioteca openpyxl não está disponível.")
        return
    
    base_dn = get_base_dn(conexao)
    print("\n📊 GERANDO RELATÓRIO - CONTAS DESABILITADAS EM 2024")
    
    data_inicio_2024 = datetime(2024, 1, 1)
    data_fim_2024 = datetime(2025, 1, 1)
    
    # Busca todos os usuários desabilitados
    filtro = '(&(objectClass=user)(!(sAMAccountName=*$))(userAccountControl:1.2.840.113556.1.4.803:=2))'
    atributos = ['sAMAccountName', 'displayName', 'title', 'mail', 'whenCreated', 'lastLogon', 'lastLogonTimestamp']
    
    dados_usuarios = buscar_usuarios_com_paginacao(conexao, base_dn, filtro, atributos)
    
    if not dados_usuarios:
        print("❌ Nenhuma conta desabilitada encontrada.")
        return
    
    # Processa os dados e aplica filtro (desabilitados que tiveram atividade em 2024)
    usuarios_processados = []
    for entry in dados_usuarios:
        # Verifica se teve atividade em 2024
        data_ultimo_logon = None
        
        if hasattr(entry, 'lastLogon') and entry.lastLogon.value:
            try:
                data_ultimo_logon = entry.lastLogon.value.replace(tzinfo=None)
            except:
                pass
        
        if hasattr(entry, 'lastLogonTimestamp') and entry.lastLogonTimestamp.value:
            try:
                logon_timestamp = entry.lastLogonTimestamp.value.replace(tzinfo=None)
                if data_ultimo_logon:
                    data_ultimo_logon = max(data_ultimo_logon, logon_timestamp)
                else:
                    data_ultimo_logon = logon_timestamp
            except:
                pass
        
        # Só inclui se teve logon em 2024
        if data_ultimo_logon and data_inicio_2024 <= data_ultimo_logon < data_fim_2024:
            login = entry.sAMAccountName.value if hasattr(entry, 'sAMAccountName') and entry.sAMAccountName.value else 'N/A'
            nome = entry.displayName.value if hasattr(entry, 'displayName') and entry.displayName.value else login
            cargo = entry.title.value if hasattr(entry, 'title') and entry.title.value else 'Não informado'
            email = entry.mail.value if hasattr(entry, 'mail') and entry.mail.value else 'N/A'
            
            # Data de criação
            data_criacao = 'N/A'
            if hasattr(entry, 'whenCreated') and entry.whenCreated.value:
                try:
                    created_dt = entry.whenCreated.value.replace(tzinfo=None)
                    data_criacao = created_dt.strftime('%d/%m/%Y')
                except:
                    data_criacao = 'Erro ao converter'
            
            usuarios_processados.append({
                'Login': login,
                'Nome': nome,
                'E-mail': email,
                'Cargo': cargo,
                'Status': 'Inativo',
                'Data de Criação': data_criacao,
                'Último Logon': data_ultimo_logon.strftime('%d/%m/%Y')
            })
    
    # Ordena por nome
    usuarios_processados.sort(key=lambda x: x['Nome'] or 'ZZZ')
    
    # Gera a planilha
    nome_arquivo = f"Contas_Desabilitadas_2024_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    colunas = ['Login', 'Nome', 'E-mail', 'Cargo', 'Status', 'Data de Criação', 'Último Logon']
    gerar_planilha(usuarios_processados, nome_arquivo, "CONTAS DESABILITADAS EM 2024", colunas)

def gerar_relacao_emails(conexao):
    """Gera relatório com todos os e-mails: Nome, E-mail e Cargo"""
    if not OPENPYXL_DISPONIVEL:
        print("❌ ERRO: Biblioteca openpyxl não está disponível.")
        return
    
    base_dn = get_base_dn(conexao)
    print("\n📊 GERANDO RELATÓRIO - RELAÇÃO DE E-MAILS")
    
    # Busca todos os usuários que têm e-mail
    filtro = '(&(objectClass=user)(!(sAMAccountName=*$))(mail=*))'
    atributos = ['sAMAccountName', 'displayName', 'title', 'mail']
    
    dados_usuarios = buscar_usuarios_com_paginacao(conexao, base_dn, filtro, atributos)
    
    if not dados_usuarios:
        print("❌ Nenhum usuário com e-mail encontrado.")
        return
    
    # Processa os dados
    usuarios_processados = []
    for entry in dados_usuarios:
        nome = entry.displayName.value if hasattr(entry, 'displayName') and entry.displayName.value else 'N/A'
        cargo = entry.title.value if hasattr(entry, 'title') and entry.title.value else 'Não informado'
        email = entry.mail.value if hasattr(entry, 'mail') and entry.mail.value else 'N/A'
        
        if email != 'N/A':  # Só inclui se tem e-mail válido
            usuarios_processados.append({
                'Nome': nome,
                'E-mail': email,
                'Cargo': cargo
            })
    
    # Ordena por nome
    usuarios_processados.sort(key=lambda x: x['Nome'] or 'ZZZ')
    
    # Gera a planilha
    nome_arquivo = f"Relacao_Emails_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    colunas = ['Nome', 'E-mail', 'Cargo']
    gerar_planilha(usuarios_processados, nome_arquivo, "RELAÇÃO DE E-MAILS", colunas)

def buscar_usuarios_com_paginacao(conexao, base_dn, filtro, atributos):
    """Função auxiliar para buscar usuários com paginação"""
    print("🔍 Buscando usuários no Active Directory...")
    
    todas_entradas = []
    cookie = None
    pagina = 0
    
    while True:
        pagina += 1
        print(f"   Buscando página {pagina}...")
        
        # Busca com paginação
        conexao.search(
            base_dn,
            filtro,
            attributes=atributos,
            paged_size=1000,
            paged_cookie=cookie,
            search_scope='SUBTREE',
            time_limit=0,
            size_limit=0
        )
        
        if not conexao.entries:
            print(f"   Nenhum resultado na página {pagina}")
            break
        
        # Adiciona os resultados à lista
        todas_entradas.extend(list(conexao.entries))
        print(f"   ✓ Página {pagina}: {len(conexao.entries)} usuários | Total: {len(todas_entradas)}")
        
        # Verifica se há mais páginas
        try:
            controls = conexao.result.get('controls', {})
            if isinstance(controls, dict):
                paged_control = controls.get('1.2.840.113556.1.4.319', {})
                if isinstance(paged_control, dict):
                    cookie = paged_control.get('value', {}).get('cookie')
                else:
                    cookie = None
            else:
                cookie = None
        except:
            cookie = None
            
        if not cookie:
            print(f"   Busca concluída após {pagina} páginas")
            break
    
    print(f"✅ Total de usuários encontrados: {len(todas_entradas)}")
    return todas_entradas

def gerar_planilha(dados, nome_arquivo, titulo, colunas):
    """Função auxiliar para gerar planilhas Excel"""
    print(f"📝 Criando planilha: {nome_arquivo}")
    
    try:
        # Cria workbook e worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = titulo.replace(' ', '_')[:30]  # Limita o nome da aba
        
        # Adiciona informações do cabeçalho
        ws['A1'] = titulo
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M:%S")}'
        ws['A3'] = f'Total de registros: {len(dados)}'
        
        # Linha vazia
        ws.append([''])
        
        # Define cabeçalhos
        ws.append(colunas)
        
        # Formata cabeçalho
        header_row = ws.max_row
        for col in range(1, len(colunas) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Adiciona dados
        for usuario in dados:
            linha = []
            for coluna in colunas:
                linha.append(usuario.get(coluna, 'N/A'))
            ws.append(linha)
        
        # Ajusta largura das colunas
        for col in range(1, len(colunas) + 1):
            ws.column_dimensions[chr(64 + col)].width = 25
        
        # Salva o arquivo
        wb.save(nome_arquivo)
        
        print(f"✅ Relatório gerado com sucesso!")
        print(f"   📄 Arquivo: {nome_arquivo}")
        print(f"   📊 Total de registros: {len(dados)}")
        
        # Abre o arquivo automaticamente
        os.startfile(nome_arquivo)
        print("✓ Arquivo aberto")
        
    except Exception as e:
        print(f"❌ Erro ao criar planilha: {e}")

def gerar_auditoria_2024(conexao):
    """Gera relatório de auditoria 2024 com critérios específicos"""
    if not OPENPYXL_DISPONIVEL:
        print("❌ ERRO: Biblioteca openpyxl não está disponível.")
        print("Para usar esta função, instale com: pip install openpyxl")
        return
    
    base_dn = get_base_dn(conexao)
    
    print("\n📊 GERANDO AUDITORIA 2024")
    print("Aplicando critérios específicos de data e status...")
    
    # Define as datas de corte
    data_inicio_2024 = datetime(2024, 1, 1)
    data_fim_periodo = datetime(2025, 1, 1)
    
    # Busca TODOS os usuários (ativos e inativos, mas não computadores)
    filtro = '(&(objectClass=user)(!(sAMAccountName=*$)))'
    atributos = ['sAMAccountName', 'displayName', 'title', 'userAccountControl', 'whenCreated', 'mail', 'lastLogon', 'lastLogonTimestamp', 'accountExpires']
    
    try:
        # Busca com paginação para lidar com muitos usuários
        print("🔍 Buscando TODOS os usuários do Active Directory...")
        
        todas_entradas = []
        cookie = None
        pagina = 0
        
        while True:
            pagina += 1
            print(f"   Buscando página {pagina}...")
            
            # Busca com paginação
            conexao.search(
                base_dn,
                filtro,
                attributes=atributos,
                paged_size=1000,
                paged_cookie=cookie,
                search_scope='SUBTREE',
                time_limit=0,
                size_limit=0
            )
            
            if not conexao.entries:
                print(f"   Nenhum resultado na página {pagina}")
                break
            
            # Adiciona os resultados à lista
            todas_entradas.extend(list(conexao.entries))
            print(f"   ✓ Página {pagina}: {len(conexao.entries)} usuários | Total: {len(todas_entradas)}")
            
            # Verifica se há mais páginas
            try:
                controls = conexao.result.get('controls', {})
                if isinstance(controls, dict):
                    paged_control = controls.get('1.2.840.113556.1.4.319', {})
                    if isinstance(paged_control, dict):
                        cookie = paged_control.get('value', {}).get('cookie')
                    else:
                        cookie = None
                else:
                    cookie = None
            except:
                cookie = None
                
            if not cookie:
                print(f"   Busca concluída após {pagina} páginas")
                break
        
        if not todas_entradas:
            print("❌ Nenhum usuário encontrado.")
            return
        
        print(f"✅ Total de usuários encontrados: {len(todas_entradas)}")
        print("� Organizando usuários em ordem alfabética...")
        
        # Ordena as entradas por sAMAccountName em ordem alfabética
        todas_entradas.sort(key=lambda x: x.sAMAccountName.value.lower() if hasattr(x, 'sAMAccountName') and x.sAMAccountName.value else 'zzz')
        
        print("�🔄 Aplicando critérios de auditoria 2024...")
        
        # Prepara dados para a planilha
        dados_usuarios = []
        usuarios_processados = 0
        usuarios_incluidos = 0
        
        for entry in todas_entradas:
            try:
                usuarios_processados += 1
                
                # Mostra progresso a cada 500 usuários
                if usuarios_processados % 500 == 0:
                    print(f"   📊 Processados: {usuarios_processados}/{len(todas_entradas)} usuários | Incluídos: {usuarios_incluidos}")
                
                login = entry.sAMAccountName.value if hasattr(entry, 'sAMAccountName') and entry.sAMAccountName.value else 'N/A'
                nome = entry.displayName.value if hasattr(entry, 'displayName') and entry.displayName.value else login
                
                # Verifica se o campo title existe e tem valor
                cargo = 'Não informado'
                if hasattr(entry, 'title') and entry.title.value:
                    cargo = entry.title.value
                
                # Verifica se o usuário está ativo
                user_account_control = int(entry.userAccountControl.value) if hasattr(entry, 'userAccountControl') and entry.userAccountControl.value else 0
                is_disabled = bool(user_account_control & 0x0002)  # Flag ACCOUNTDISABLE
                status = 'Inativo' if is_disabled else 'Ativo'
                
                # Converte data de criação
                data_criacao = 'N/A'
                created_dt = None
                if hasattr(entry, 'whenCreated') and entry.whenCreated.value:
                    try:
                        created_raw = entry.whenCreated.value
                        if isinstance(created_raw, datetime):
                            # Já é um objeto datetime, remove timezone se presente
                            created_dt = created_raw.replace(tzinfo=None)
                            data_criacao = created_dt.strftime('%d/%m/%Y')
                        else:
                            # Formato não esperado
                            data_criacao = 'Formato inválido'
                    except (ValueError, TypeError):
                        data_criacao = 'Erro ao converter'
                
                # Extrai e-mail
                email = 'N/A'
                if hasattr(entry, 'mail') and entry.mail.value:
                    email = entry.mail.value
                
                # Converte data de expiração da conta do AD
                data_expiracao_ad = 'Nunca expira'
                account_expires_dt = None
                tem_data_expiracao_valida = False
                
                if hasattr(entry, 'accountExpires') and entry.accountExpires.value:
                    try:
                        expires_raw = entry.accountExpires.value
                        
                        # Caso 1: Valor convertido incorretamente para 1601 (valor 0 no AD = nunca expira)
                        if isinstance(expires_raw, datetime) and expires_raw.year == 1601:
                            # Verifica se realmente é 0 acessando o valor bruto
                            try:
                                raw_value = entry.accountExpires.raw_values[0] if entry.accountExpires.raw_values else None
                                if raw_value:
                                    if isinstance(raw_value, bytes):
                                        expires_ticks = int(raw_value.decode('utf-8'))
                                    else:
                                        expires_ticks = int(raw_value)
                                    
                                    if expires_ticks == 0:
                                        # Valor 0 = nunca expira
                                        data_expiracao_ad = 'Nunca expira'
                                        tem_data_expiracao_valida = False
                                    elif expires_ticks != 9223372036854775807:
                                        # Valor válido, converte
                                        temp_dt = datetime(1601, 1, 1) + timedelta(seconds=expires_ticks/10000000)
                                        if temp_dt.year == 9999:
                                            data_expiracao_ad = 'Nunca expira'
                                            tem_data_expiracao_valida = False
                                        else:
                                            data_expiracao_ad = temp_dt.strftime('%d/%m/%Y')
                                            account_expires_dt = temp_dt
                                            tem_data_expiracao_valida = True
                                    else:
                                        # Valor máximo = nunca expira
                                        data_expiracao_ad = 'Nunca expira'
                                        tem_data_expiracao_valida = False
                                else:
                                    data_expiracao_ad = 'Nunca expira'
                                    tem_data_expiracao_valida = False
                            except (ValueError, TypeError, AttributeError):
                                data_expiracao_ad = 'Nunca expira'
                                tem_data_expiracao_valida = False
                        
                        # Caso 2: Datetime já convertido corretamente
                        elif isinstance(expires_raw, datetime):
                            account_expires_dt = expires_raw.replace(tzinfo=None)
                            if account_expires_dt.year == 9999:
                                data_expiracao_ad = 'Nunca expira'
                                tem_data_expiracao_valida = False
                            else:
                                data_expiracao_ad = account_expires_dt.strftime('%d/%m/%Y')
                                tem_data_expiracao_valida = True
                        
                        # Caso 3: Valor em ticks do Windows
                        elif isinstance(expires_raw, (int, str)):
                            expires_ticks = int(expires_raw)
                            if expires_ticks == 0:
                                data_expiracao_ad = 'Nunca expira'
                                tem_data_expiracao_valida = False
                            elif expires_ticks != 9223372036854775807:
                                temp_dt = datetime(1601, 1, 1) + timedelta(seconds=expires_ticks/10000000)
                                if temp_dt.year == 9999:
                                    data_expiracao_ad = 'Nunca expira'
                                    tem_data_expiracao_valida = False
                                else:
                                    data_expiracao_ad = temp_dt.strftime('%d/%m/%Y')
                                    account_expires_dt = temp_dt
                                    tem_data_expiracao_valida = True
                            else:
                                data_expiracao_ad = 'Nunca expira'
                                tem_data_expiracao_valida = False
                        else:
                            data_expiracao_ad = 'Nunca expira'
                            tem_data_expiracao_valida = False
                    except (ValueError, TypeError, OverflowError):
                        data_expiracao_ad = 'Nunca expira'
                        tem_data_expiracao_valida = False
                
                # Obtém datas de último logon
                last_logon_dt = None
                last_logon_timestamp_dt = None
                
                # Processa lastLogon
                if hasattr(entry, 'lastLogon') and entry.lastLogon.value:
                    try:
                        last_logon_raw = entry.lastLogon.value
                        if isinstance(last_logon_raw, datetime):
                            # Já é um objeto datetime, remove timezone se presente
                            last_logon_dt = last_logon_raw.replace(tzinfo=None)
                        elif isinstance(last_logon_raw, (int, str)):
                            # Valor em ticks do Windows
                            last_logon_ticks = int(last_logon_raw)
                            if last_logon_ticks > 0:
                                temp_dt = datetime(1601, 1, 1) + timedelta(seconds=last_logon_ticks/10000000)
                                if temp_dt.year != 1601 and temp_dt.year != 9999:
                                    last_logon_dt = temp_dt
                    except (ValueError, TypeError, OverflowError):
                        pass
                
                # Processa lastLogonTimestamp
                if hasattr(entry, 'lastLogonTimestamp') and entry.lastLogonTimestamp.value:
                    try:
                        last_logon_timestamp_raw = entry.lastLogonTimestamp.value
                        if isinstance(last_logon_timestamp_raw, datetime):
                            # Já é um objeto datetime, remove timezone se presente
                            last_logon_timestamp_dt = last_logon_timestamp_raw.replace(tzinfo=None)
                        elif isinstance(last_logon_timestamp_raw, (int, str)):
                            # Valor em ticks do Windows
                            last_logon_timestamp_ticks = int(last_logon_timestamp_raw)
                            if last_logon_timestamp_ticks > 0:
                                temp_dt = datetime(1601, 1, 1) + timedelta(seconds=last_logon_timestamp_ticks/10000000)
                                if temp_dt.year != 1601 and temp_dt.year != 9999:
                                    last_logon_timestamp_dt = temp_dt
                    except (ValueError, TypeError, OverflowError):
                        pass
                
                # Obtém a data mais recente entre lastLogon e lastLogonTimestamp
                data_ultimo_logon_mais_recente = None
                if last_logon_dt and last_logon_timestamp_dt:
                    data_ultimo_logon_mais_recente = max(last_logon_dt, last_logon_timestamp_dt)
                elif last_logon_dt:
                    data_ultimo_logon_mais_recente = last_logon_dt
                elif last_logon_timestamp_dt:
                    data_ultimo_logon_mais_recente = last_logon_timestamp_dt
                
                # APLICAÇÃO DOS CRITÉRIOS DE AUDITORIA 2024
                incluir_usuario = False
                motivo_exclusao = ""
                
                # REGRA 1: Todas as contas atualmente ativas devem ser exibidas
                if status == 'Ativo':
                    incluir_usuario = True
                    motivo_exclusao = "Conta ativa - incluída automaticamente"
                
                # REGRA 2: Para contas inativas, aplicar critérios específicos
                elif status == 'Inativo':
                    # Critério A: Data de criação >= 01/01/2024
                    if created_dt and created_dt >= data_inicio_2024 and created_dt < datetime(2025, 1, 1):
                        incluir_usuario = True
                        motivo_exclusao = f"Conta inativa criada em 2024 ou posterior ({created_dt.strftime('%d/%m/%Y')})"
                    
                    # Critério B: Último logon >= 01/01/2024 e < 01/01/2025
                    if not incluir_usuario and data_ultimo_logon_mais_recente:
                        if data_ultimo_logon_mais_recente >= data_inicio_2024 and data_ultimo_logon_mais_recente < datetime(2025, 1, 1):
                            incluir_usuario = True
                            motivo_exclusao = f"Conta inativa com último logon em 2024 ({data_ultimo_logon_mais_recente.strftime('%d/%m/%Y')})"
                    
                    # Critério C: Contas sem data de expiração válida (nunca expira)
                    # Só incluir se tiver login em 2024
                    if not incluir_usuario and not tem_data_expiracao_valida:
                        if data_ultimo_logon_mais_recente and data_ultimo_logon_mais_recente >= data_inicio_2024 and data_ultimo_logon_mais_recente < datetime(2025, 1, 1):
                            incluir_usuario = True
                            motivo_exclusao = f"Conta inativa sem data de expiração, mas com logon em 2024 ({data_ultimo_logon_mais_recente.strftime('%d/%m/%Y')})"
                        else:
                            motivo_exclusao = "Conta inativa sem data de expiração válida e sem logon em 2024 - excluída"
                    
                    # Critério D: Exclusão de contas criadas em 2025 ou posterior
                    if incluir_usuario and created_dt >= datetime(2025, 1, 1):
                        incluir_usuario = False
                        motivo_exclusao = f"Conta inativa criada em 2025 ou posterior ({created_dt.strftime('%d/%m/%Y')}) - excluída"
                    
                    # Critério E: Exclusão de contas com login em 2025 ou posterior
                    if incluir_usuario and data_ultimo_logon_mais_recente >= datetime(2025, 1, 1):
                        incluir_usuario = False
                        motivo_exclusao = f"Conta inativa com último logon em 2025 ou posterior ({data_ultimo_logon_mais_recente.strftime('%d/%m/%Y')}) - excluída"
                    
                    # Critério F: Exclusão de contas sem atividade em 2024
                    if incluir_usuario and (data_ultimo_logon_mais_recente < data_inicio_2024 or data_ultimo_logon_mais_recente > data_fim_periodo):
                        incluir_usuario = False
                        motivo_exclusao = "Conta inativa sem atividade em 2024 - excluída"

                    # Não atende nenhum critério
                    if not incluir_usuario:
                        if data_ultimo_logon_mais_recente:
                            if data_ultimo_logon_mais_recente < data_inicio_2024:
                                motivo_exclusao = f"Último logon anterior a 2024 ({data_ultimo_logon_mais_recente.strftime('%d/%m/%Y')}) - excluída"
                        else:
                            motivo_exclusao = "Conta inativa sem atividade em 2024 - excluída"

                # Debug: mostra informações detalhadas para os primeiros 350 usuários
                if usuarios_processados <= 350:
                    print(f"  DEBUG {usuarios_processados} - {login}:")
                    print(f"    Status: {status}")
                    print(f"    Criação: {created_dt.strftime('%d/%m/%Y') if created_dt else 'N/A'}")
                    print(f"    Último logon: {data_ultimo_logon_mais_recente.strftime('%d/%m/%Y') if data_ultimo_logon_mais_recente else 'Nunca'}")
                    print(f"    Expiração AD: {data_expiracao_ad}")
                    print(f"    Expiração DT: {account_expires_dt.strftime('%d/%m/%Y') if account_expires_dt else 'Nunca expira'}")
                    print(f"    Tem data válida: {'✅ SIM' if tem_data_expiracao_valida else '❌ NÃO'}")
                    print(f"    Incluir: {'✅ SIM' if incluir_usuario else '❌ NÃO'}")
                    print(f"    Motivo: {motivo_exclusao}")
                    print(f"    ---")
                
                # Se deve incluir o usuário, adiciona aos dados
                if incluir_usuario:
                    # Calcula a data de expiração personalizada para a planilha
                    data_expiracao_personalizada = data_expiracao_ad
                    
                    # Se há data de último logon mais recente no período, usa ela como referência
                    if data_ultimo_logon_mais_recente and data_inicio_2024 <= data_ultimo_logon_mais_recente <= data_fim_periodo:
                        data_expiracao_personalizada = data_ultimo_logon_mais_recente.strftime('%d/%m/%Y')
                    
                    dados_usuarios.append({
                        'Login': login,
                        'Nome': nome,
                        'E-mail': email,
                        'Cargo': cargo,
                        'Status': status,
                        'Data de Criação': data_criacao,
                        'Data de Expiração': data_expiracao_personalizada
                    })
                    usuarios_incluidos += 1
                    
            except Exception as e:
                print(f"⚠ Erro ao processar usuário {login}: {e}")
                continue
        
        print(f"✓ Processados: {usuarios_processados} usuários")
        print(f"✓ Incluídos na auditoria: {usuarios_incluidos} usuários")
        
        if not dados_usuarios:
            print("❌ Nenhum usuário atende aos critérios de auditoria 2024.")
            return
        
        # Ordena por nome
        dados_usuarios.sort(key=lambda x: x['Nome'] or 'ZZZ')
        
        # Cria a planilha
        nome_arquivo = f"Auditoria_2024_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        print(f"📝 Criando planilha: {nome_arquivo}")
        
        try:
            # Cria workbook e worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Auditoria 2024"
            
            # Adiciona informações do cabeçalho
            ws['A1'] = 'AUDITORIA 2024 - ACTIVE DIRECTORY'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M:%S")}'
            ws['A3'] = f'Critérios: Contas ativas + inativas com criação/logon em 2024'
            ws['A4'] = f'Total de usuários na auditoria: {len(dados_usuarios)}'
            
            # Conta usuários ativos e inativos
            ativos = sum(1 for u in dados_usuarios if u['Status'] == 'Ativo')
            inativos = len(dados_usuarios) - ativos
            ws['A5'] = f'Usuários ativos: {ativos} | Usuários inativos: {inativos}'
            
            # Adiciona informações sobre usuários excluídos
            ws['A6'] = f'Total de usuários processados: {usuarios_processados}'
            ws['A7'] = f'Usuários excluídos da auditoria: {usuarios_processados - usuarios_incluidos}'
            
            # Linha vazia
            ws.append([''])
            
            # Define cabeçalhos
            colunas = ['Login', 'Nome', 'E-mail', 'Cargo', 'Status', 'Data de Criação', 'Data de Expiração']
            ws.append(colunas)
            
            # Formata cabeçalho
            header_row = ws.max_row
            for col in range(1, len(colunas) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Adiciona dados
            for usuario in dados_usuarios:
                ws.append([
                    usuario['Login'],
                    usuario['Nome'],
                    usuario['E-mail'],
                    usuario['Cargo'],
                    usuario['Status'],
                    usuario['Data de Criação'],
                    usuario['Data de Expiração']
                ])
            
            # Ajusta largura das colunas
            ws.column_dimensions['A'].width = 18  # Login
            ws.column_dimensions['B'].width = 35  # Nome
            ws.column_dimensions['C'].width = 30  # E-mail
            ws.column_dimensions['D'].width = 25  # Cargo
            ws.column_dimensions['E'].width = 10  # Status
            ws.column_dimensions['F'].width = 15  # Data de Criação
            ws.column_dimensions['G'].width = 15  # Data de Expiração
            
            # Salva o arquivo
            wb.save(nome_arquivo)
            
            print(f"✅ Auditoria 2024 gerada com sucesso!")
            print(f"   📄 Arquivo: {nome_arquivo}")
            print(f"   👥 Total de usuários: {len(dados_usuarios)}")
            print(f"   ✅ Usuários ativos: {ativos}")
            print(f"   ❌ Usuários inativos: {inativos}")
            print(f"   📊 Colunas: {', '.join(colunas)}")
            
            # Abre o arquivo automaticamente
            os.startfile(nome_arquivo)
            print("✓ Arquivo aberto")
            
        except Exception as e:
            print(f"❌ Erro ao criar planilha: {e}")
            
    except Exception as e:
        print(f"❌ Erro ao buscar usuários: {e}")
        print("Verifique se você tem permissões para listar usuários no AD.")

def gerar_auditoria_2024(conexao):
    """
    Gera auditoria completa 2024 com critério original
    """
    print("\n📋 Critérios aplicados:")
    print("   • Todas as contas ATIVAS são incluídas")
    print("   • Contas INATIVAS são incluídas SE:")
    print("     - Data de criação >= 01/01/2024 OU")
    print("     - Último logon >= 01/01/2024 e < 01/01/2025")
    print("   • EXCLUSÃO: contas inativas antigas sem atividade em 2024")
    
    # Buscar todos os usuários
    usuarios = buscar_usuarios_com_paginacao(conexao)
    
    # Definir datas de referência
    inicio_2024 = datetime(2024, 1, 1)
    fim_2024 = datetime(2024, 12, 31, 23, 59, 59)
    
    usuarios_incluidos = []
    
    for usuario in usuarios:
        # Obter atributos do usuário
        attrs = usuario.entry_attributes_as_dict
        
        # Verificar se a conta está ativa
        uac = attrs.get('userAccountControl', [0])[0]
        conta_ativa = not (uac & 0x00000002)  # Bit 1 = ACCOUNTDISABLE
        
        # Se a conta está ativa, incluir sempre
        if conta_ativa:
            usuarios_incluidos.append(usuario)
            continue
        
        # Para contas inativas, verificar critérios de 2024
        incluir_conta = False
        
        # Verificar data de criação
        when_created = attrs.get('whenCreated', [])
        if when_created:
            try:
                data_criacao = when_created[0]
                if isinstance(data_criacao, str):
                    data_criacao = datetime.strptime(data_criacao, '%Y-%m-%d %H:%M:%S%z')
                elif hasattr(data_criacao, 'replace'):
                    data_criacao = data_criacao.replace(tzinfo=None)
                
                if data_criacao >= inicio_2024:
                    incluir_conta = True
            except:
                pass
        
        # Verificar último logon
        if not incluir_conta:
            for attr_logon in ['lastLogon', 'lastLogonTimestamp']:
                if incluir_conta:
                    break
                
                logon_values = attrs.get(attr_logon, [])
                if logon_values:
                    try:
                        logon_timestamp = logon_values[0]
                        if isinstance(logon_timestamp, int) and logon_timestamp > 0:
                            data_logon = datetime.utcfromtimestamp(
                                (logon_timestamp - 116444736000000000) / 10000000
                            )
                            if inicio_2024 <= data_logon <= fim_2024:
                                incluir_conta = True
                                break
                    except:
                        pass
        
        if incluir_conta:
            usuarios_incluidos.append(usuario)
    
    # Gerar planilha
    filename = f"Auditoria_2024_Completa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    gerar_planilha(usuarios_incluidos, filename)
    
    print(f"\n✅ Auditoria 2024 gerada com sucesso!")
    print(f"📄 Arquivo: {filename}")
    print(f"👥 Total de usuários incluídos: {len(usuarios_incluidos)}")

def menu():
    print("\n" + "="*60)
    print("        AUDITORIA ACTIVE DIRECTORY - 2024")
    print("="*60)
    
    try:
        conexao = get_conexao()
        print("\n✅ Conectado ao Active Directory com sucesso!")
        
        while True:
            print("\n� MENU DE RELATÓRIOS")
            print("="*40)
            print("1️⃣  Contas de usuário ativas")
            print("2️⃣  Contas desabilitadas a partir de 01/04/2024")
            print("3️⃣  Contas de usuários criadas somente em 2024")
            print("4️⃣  Contas de usuários desabilitadas somente em 2024")
            print("5️⃣  Relação de todos os e-mails (Nome, E-mail, Cargo)")
            print("6️⃣  Auditoria 2024 (Critério completo original)")
            print("0️⃣  Sair")
            print("="*40)
            
            try:
                opcao = input("\n🔍 Escolha uma opção (0-6): ").strip()
                
                if opcao == '0':
                    print("\n👋 Encerrando o programa...")
                    break
                elif opcao == '1':
                    print("\n🔄 Gerando relatório de contas ativas...")
                    gerar_contas_ativas(conexao)
                elif opcao == '2':
                    print("\n🔄 Gerando relatório de contas desabilitadas desde abril/2024...")
                    gerar_contas_desabilitadas_desde_abril(conexao)
                elif opcao == '3':
                    print("\n🔄 Gerando relatório de contas criadas em 2024...")
                    gerar_contas_criadas_em_2024(conexao)
                elif opcao == '4':
                    print("\n🔄 Gerando relatório de contas desabilitadas em 2024...")
                    gerar_contas_desabilitadas_em_2024(conexao)
                elif opcao == '5':
                    print("\n🔄 Gerando relação de e-mails...")
                    gerar_relacao_emails(conexao)
                elif opcao == '6':
                    print("\n🔄 Gerando auditoria 2024 (critério completo)...")
                    gerar_auditoria_2024(conexao)
                else:
                    print("❌ Opção inválida! Escolha uma opção entre 0 e 6.")
                    continue
                
                print("\n" + "="*60)
                input("Pressione Enter para continuar...")
                
            except KeyboardInterrupt:
                print("\n\n👋 Programa interrompido pelo usuário.")
                break
            except Exception as e:
                print(f"\n❌ Erro ao executar opção: {e}")
                input("Pressione Enter para continuar...")
        
    except Exception as e:
        print(f"❌ Erro fatal: {e}")
        input("Pressione Enter para encerrar...")
    finally:
        print("\n🔚 Programa finalizado.")

if __name__ == "__main__":
    menu()
